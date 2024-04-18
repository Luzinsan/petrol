from openpyxl.utils import range_boundaries
from openpyxl import load_workbook
import datetime
import os

NUM_MERGE=3
DATA = './data'
RAW_DATA =  os.path.join(DATA, 'raw/')
PATH = os.path.join(RAW_DATA, 'Автоматизированные_измерения.xlsx')
print(f"Чтение данных: {PATH}")
wb = load_workbook(PATH)
sheet = wb['Лист1']


############Flatten names into single-row header##################
print("Размерживание...")
merge_list = []
for merge in sheet.merged_cells.ranges:
    merge_list.append(merge)
for merge in merge_list:
    # Находим границы смерженных ячеек
    min_col, min_row, max_col, max_row = range_boundaries(merge.coord)
    # Запоминаем значение смёрженной ячейки, чтобы потом его продублировать
    top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
    sheet.unmerge_cells(str(merge))
    for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
        for cell in row: # дублируем значение
            cell.value = top_left_cell_value

print('Расспаршивание заголовка...')
# генератор добавляет символ '|' перед значением ячейки, если оно не null (null+str недопустим)
generator = lambda x: f"|{x.value}" if x.value else ""
# здесь рассчитано, что 3 (NUM_MERGE) строки (заголовка) объединяются в однострочный заголовок
# можно запрограммировать до параметра, но мне лень
for cols in sheet.iter_cols(min_row=NUM_MERGE, max_row=NUM_MERGE, max_col=sheet.max_column):
    for cell in cols:
        cell.value = (
            generator(sheet.cell(row=1,column=cell.column))+
            generator(sheet.cell(row=2,column=cell.column))+
            generator(cell)
            )[1:] # убираем лишний '|'
        # print(cell.value)

sheet.delete_rows(1,2) # размержили и распарсили заголовки, больше не нужны

####################### Split date and time ######################
# В этом блоке будем дублировать значение даты (она написана только в ключевых моментах)
# А также разделяем время в отдельный столбик
# Пример:
# Было             | Станет:
# Дата/Время       | Дата         Время
# 01.01.2021  2:00 | 01.01.2021    2:00
#             4:00 | 01.01.2021    4:00
#             6:00 | 01.01.2021    6:00
print('Расспаршивание "Дата/Время"...')
sheet.insert_cols(1)
sheet.cell(row=1, column=1).value = "Дата"
sheet.cell(row=1, column=2).value = "Время"
date = None

for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2):
    for cell in row:
        value = cell.value 
        if isinstance(value, datetime.datetime): # Если это запись формата DD.MM.YY HH:00
            # И она не выпадает на полночь (00:00 в datetime воспринимается как datetime.datetime)
            if str(value) != '1900-01-01 00:00:00': 
                date = value.date() # фиксируем новый день
            # оставляем только час (01.01.2021  2:00 => 2:00)
            cell.value = value.time() 
            cell.number_format = 'H:mm'
        # записываем дату в соседнюю колонку
        sheet.cell(row=cell.row, column=1).value = date 
        
print("Сохраняем...")
wb.save(os.path.join(DATA, 'processed.xlsx'))
print("The end!")
